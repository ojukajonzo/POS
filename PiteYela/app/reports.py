"""
Reports and financial analysis module for Alcohol POS System.
Admin-only sales reports, profit/loss analysis, and exports.
"""
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple
from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QTableWidget,
                             QTableWidgetItem, QPushButton, QLabel, QDateEdit,
                             QComboBox, QGroupBox, QFormLayout, QMessageBox, QFileDialog)
from PyQt6.QtCore import QDate, Qt
from app.database import get_connection
from app.auth import is_admin
from app.utils import format_currency, format_date, get_date_range, show_error_dialog, show_info_dialog
from datetime import datetime

class ReportsWindow(QWidget):
    """Reports and financial analysis window (Admin only)."""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        if not is_admin():
            show_error_dialog(self, "Access Denied", "Admin access required")
            return
        
        self.setWindowTitle("PiteYelaHouseofWine_POS - Sales Reports & Financial Analysis")
        self.resize(1200, 700)
        self.setStyleSheet("background-color: white;")
        self.setWindowModality(Qt.WindowModality.WindowModal)
        
        layout = QVBoxLayout()
        
        # Back button
        back_layout = QHBoxLayout()
        self.back_btn = QPushButton("← Back")
        self.back_btn.clicked.connect(self.close)
        back_layout.addWidget(self.back_btn)
        back_layout.addStretch()
        layout.addLayout(back_layout)
        
        # Filter section
        filter_group = QGroupBox("Date Filter")
        filter_group.setStyleSheet("""
            QGroupBox {
                border: 2px solid #ccc;
                border-radius: 0px;
                margin-top: 10px;
                padding-top: 15px;
                background-color: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                background-color: white;
            }
        """)
        filter_layout = QFormLayout()
        
        self.period_combo = QComboBox()
        self.period_combo.addItems(["Day", "Week", "Month", "Custom Range"])
        self.period_combo.currentTextChanged.connect(self.on_period_changed)
        filter_layout.addRow("Period:", self.period_combo)
        
        self.start_date = QDateEdit()
        self.start_date.setDate(QDate.currentDate())
        self.start_date.setCalendarPopup(True)
        filter_layout.addRow("Start Date:", self.start_date)
        
        self.end_date = QDateEdit()
        self.end_date.setDate(QDate.currentDate())
        self.end_date.setCalendarPopup(True)
        filter_layout.addRow("End Date:", self.end_date)
        
        self.generate_btn = QPushButton("Generate Report")
        self.generate_btn.clicked.connect(self.generate_report)
        filter_layout.addRow("", self.generate_btn)
        
        # Export buttons
        export_layout = QHBoxLayout()
        self.export_pdf_btn = QPushButton("Export to PDF")
        self.export_pdf_btn.clicked.connect(self.export_to_pdf)
        self.export_excel_btn = QPushButton("Export to Excel")
        self.export_excel_btn.clicked.connect(self.export_to_excel)
        
        button_style = """
            QPushButton {
                background-color: white;
                color: black;
                border: 2px solid #000000;
                padding: 8px 16px;
                border-radius: 0px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #f0f0f0;
                border: 2px solid #000000;
            }
        """
        self.generate_btn.setStyleSheet(button_style)
        self.export_pdf_btn.setStyleSheet(button_style)
        self.export_excel_btn.setStyleSheet(button_style)
        
        export_layout.addWidget(self.export_pdf_btn)
        export_layout.addWidget(self.export_excel_btn)
        filter_layout.addRow("Export:", export_layout)
        
        filter_group.setLayout(filter_layout)
        layout.addWidget(filter_group)
        
        # Summary section
        summary_group = QGroupBox("Summary")
        summary_group.setStyleSheet("""
            QGroupBox {
                border: 2px solid #ccc;
                border-radius: 0px;
                margin-top: 10px;
                padding-top: 15px;
                background-color: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                background-color: white;
            }
        """)
        summary_layout = QHBoxLayout()
        
        self.total_sales_label = QLabel("Total Sales: UGX 0")
        self.total_transactions_label = QLabel("Transactions: 0")
        self.total_cost_label = QLabel("Total Cost: UGX 0")
        self.total_profit_label = QLabel("Total Profit: UGX 0")
        
        summary_layout.addWidget(self.total_sales_label)
        summary_layout.addWidget(self.total_transactions_label)
        summary_layout.addWidget(self.total_cost_label)
        summary_layout.addWidget(self.total_profit_label)
        summary_layout.addStretch()
        
        summary_group.setLayout(summary_layout)
        layout.addWidget(summary_group)
        
        # Sales table
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels([
            "Sale ID", "Date", "Cashier", "Items", "Total", "Profit"
        ])
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setStyleSheet("""
            QTableWidget {
                border: 2px solid #ccc;
                gridline-color: #e0e0e0;
                background-color: white;
                selection-background-color: #4CAF50;
            }
            QTableWidget::item:selected {
                background-color: #4CAF50;
                color: white;
            }
        """)
        layout.addWidget(self.table)
        
        # Cashier summary table
        cashier_group = QGroupBox("Sales by Cashier")
        cashier_group.setStyleSheet("""
            QGroupBox {
                border: 2px solid #ccc;
                border-radius: 0px;
                margin-top: 10px;
                padding-top: 15px;
                background-color: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                background-color: white;
            }
        """)
        cashier_layout = QVBoxLayout()
        
        self.cashier_table = QTableWidget()
        self.cashier_table.setColumnCount(3)
        self.cashier_table.setHorizontalHeaderLabels([
            "Cashier", "Transactions", "Total Sales"
        ])
        self.cashier_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.cashier_table.setStyleSheet("""
            QTableWidget {
                border: 2px solid #ccc;
                gridline-color: #e0e0e0;
                background-color: white;
                selection-background-color: #4CAF50;
            }
            QTableWidget::item:selected {
                background-color: #4CAF50;
                color: white;
            }
        """)
        cashier_layout.addWidget(self.cashier_table)
        
        cashier_group.setLayout(cashier_layout)
        layout.addWidget(cashier_group)
        
        self.setLayout(layout)
        
        # Generate initial report
        self.generate_report()
    
    def on_period_changed(self, period: str):
        """Update date fields based on period selection."""
        today = QDate.currentDate()
        
        if period == "Day":
            self.start_date.setDate(today)
            self.end_date.setDate(today)
        elif period == "Week":
            # Start of week (Monday)
            days_since_monday = today.dayOfWeek() - 1
            start = today.addDays(-days_since_monday)
            self.start_date.setDate(start)
            self.end_date.setDate(today)
        elif period == "Month":
            # Start of month
            start = QDate(today.year(), today.month(), 1)
            self.start_date.setDate(start)
            self.end_date.setDate(today)
        # Custom Range: user can set dates manually
    
    def get_date_range_from_ui(self) -> Tuple[datetime, datetime]:
        """Get date range from UI."""
        start_qdate = self.start_date.date()
        end_qdate = self.end_date.date()
        
        start_dt = datetime(start_qdate.year(), start_qdate.month(), start_qdate.day(), 0, 0, 0)
        end_dt = datetime(end_qdate.year(), end_qdate.month(), end_qdate.day(), 23, 59, 59)
        
        return start_dt, end_dt
    
    def generate_report(self):
        """Generate sales report based on selected filters."""
        start_dt, end_dt = self.get_date_range_from_ui()
        
        conn = get_connection()
        try:
            # Get sales in date range
            cursor = conn.cursor()
            cursor.execute("""
                SELECT s.id, s.sale_date, s.cashier_name, s.grand_total,
                       COUNT(si.id) as item_count
                FROM sales s
                LEFT JOIN sale_items si ON s.id = si.sale_id
                WHERE DATE(s.sale_date) BETWEEN DATE(?) AND DATE(?)
                GROUP BY s.id
                ORDER BY s.sale_date DESC
            """, (start_dt.isoformat(), end_dt.isoformat()))
            
            sales = cursor.fetchall()
            
            # Calculate profit for each sale
            total_sales = 0.0
            total_cost = 0.0
            total_profit = 0.0
            
            self.table.setRowCount(len(sales))
            for row_idx, sale in enumerate(sales):
                sale_id = sale['id']
                
                # Get sale items to calculate profit
                cursor.execute("""
                    SELECT si.product_id, si.quantity, si.unit_price, si.line_total,
                           p.cost_price
                    FROM sale_items si
                    JOIN products p ON si.product_id = p.id
                    WHERE si.sale_id = ?
                """, (sale_id,))
                
                items = cursor.fetchall()
                sale_profit = 0.0
                for item in items:
                    cost = item['cost_price'] * item['quantity']
                    sale_profit += item['line_total'] - cost
                
                total_sales += sale['grand_total']
                total_cost += sum(item['cost_price'] * item['quantity'] for item in items)
                total_profit += sale_profit
                
                # Populate table
                self.table.setItem(row_idx, 0, QTableWidgetItem(str(sale['id'])))
                self.table.setItem(row_idx, 1, QTableWidgetItem(sale['sale_date']))
                self.table.setItem(row_idx, 2, QTableWidgetItem(sale['cashier_name']))
                self.table.setItem(row_idx, 3, QTableWidgetItem(str(sale['item_count'])))
                self.table.setItem(row_idx, 4, QTableWidgetItem(format_currency(sale['grand_total'])))
                self.table.setItem(row_idx, 5, QTableWidgetItem(format_currency(sale_profit)))
            
            self.table.resizeColumnsToContents()
            
            # Update summary
            self.total_sales_label.setText(f"Total Sales: {format_currency(total_sales)}")
            self.total_transactions_label.setText(f"Transactions: {len(sales)}")
            self.total_cost_label.setText(f"Total Cost: {format_currency(total_cost)}")
            self.total_profit_label.setText(f"Total Profit: {format_currency(total_profit)}")
            
            # Cashier summary
            cursor.execute("""
                SELECT cashier_name, COUNT(*) as transaction_count, SUM(grand_total) as total_sales
                FROM sales
                WHERE DATE(sale_date) BETWEEN DATE(?) AND DATE(?)
                GROUP BY cashier_name
                ORDER BY total_sales DESC
            """, (start_dt.isoformat(), end_dt.isoformat()))
            
            cashiers = cursor.fetchall()
            self.cashier_table.setRowCount(len(cashiers))
            
            for row_idx, cashier in enumerate(cashiers):
                self.cashier_table.setItem(row_idx, 0, QTableWidgetItem(cashier['cashier_name']))
                self.cashier_table.setItem(row_idx, 1, QTableWidgetItem(str(cashier['transaction_count'])))
                self.cashier_table.setItem(row_idx, 2, QTableWidgetItem(format_currency(cashier['total_sales'] or 0)))
            
            self.cashier_table.resizeColumnsToContents()
            
        except Exception as e:
            show_error_dialog(self, "Error", f"Failed to generate report: {str(e)}")
        finally:
            conn.close()
    
    def export_to_pdf(self):
        """Export current report to PDF."""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            from app.config import SHOP_NAME, SHOP_LOCATION, SHOP_CONTACT
            
            filename, _ = QFileDialog.getSaveFileName(
                self, "Export to PDF", "", "PDF Files (*.pdf)"
            )
            
            if not filename:
                return
            
            if not filename.endswith('.pdf'):
                filename += '.pdf'
            
            doc = SimpleDocTemplate(filename, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()
            
            # Title and shop info
            title = Paragraph(f"<b>{SHOP_NAME} - Sales Report</b>", styles['Title'])
            elements.append(title)
            elements.append(Paragraph(str(SHOP_LOCATION or ""), styles['Normal']))
            elements.append(Paragraph(str(SHOP_CONTACT or ""), styles['Normal']))
            elements.append(Spacer(1, 12))
            
            # Date range
            start_dt, end_dt = self.get_date_range_from_ui()
            date_text = f"Period: {start_dt.strftime('%Y-%m-%d')} to {end_dt.strftime('%Y-%m-%d')}"
            elements.append(Paragraph(date_text, styles['Normal']))
            elements.append(Spacer(1, 12))
            
            # Summary
            summary_data = [
                ['Metric', 'Value'],
                ['Total Sales', self.total_sales_label.text().split(': ')[1]],
                ['Total Transactions', self.total_transactions_label.text().split(': ')[1]],
                ['Total Cost', self.total_cost_label.text().split(': ')[1]],
                ['Total Profit', self.total_profit_label.text().split(': ')[1]],
            ]
            summary_table = Table(summary_data)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 20))
            
            # Sales table
            if self.table.rowCount() > 0:
                sales_data = [['Sale ID', 'Date', 'Cashier', 'Items', 'Total', 'Profit']]
                for row in range(self.table.rowCount()):
                    row_data = []
                    for col in range(self.table.columnCount()):
                        item = self.table.item(row, col)
                        row_data.append(item.text() if item else '')
                    sales_data.append(row_data)
                
                sales_table = Table(sales_data)
                sales_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                ]))
                elements.append(Paragraph("<b>Sales Details</b>", styles['Heading2']))
                elements.append(sales_table)
            
            doc.build(elements)
            show_info_dialog(self, "Success", f"Report exported to PDF:\n{filename}")
        except ImportError:
            show_error_dialog(self, "Error", "reportlab package not installed")
        except Exception as e:
            show_error_dialog(self, "Error", f"Failed to export PDF: {str(e)}")
    
    def export_to_excel(self):
        """Export current report to Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            filename, _ = QFileDialog.getSaveFileName(
                self, "Export to Excel", "", "Excel Files (*.xlsx)"
            )
            
            if not filename:
                return
            
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Sales Report"
            
            # Title - No merge cells to avoid errors
            from app.config import SHOP_DISPLAY_NAME, SHOP_LOCATION, SHOP_CONTACT
            ws['A1'] = SHOP_DISPLAY_NAME
            ws['A1'].font = Font(bold=True, size=16)
            ws['A2'] = f"Location: {SHOP_LOCATION}"
            ws['A3'] = f"Contact: {SHOP_CONTACT}"
            ws['A4'] = "Sales Report"
            ws['A4'].font = Font(bold=True, size=14)
            
            # Date range
            start_dt, end_dt = self.get_date_range_from_ui()
            ws['A5'] = f"Period: {start_dt.strftime('%Y-%m-%d')} to {end_dt.strftime('%Y-%m-%d')}"
            
            # Summary
            row = 7
            ws[f'A{row}'] = 'Summary'
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            summary_headers = ['Metric', 'Value']
            for col, header in enumerate(summary_headers, 1):
                cell = ws.cell(row, col, header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            row += 1
            ws.cell(row, 1, 'Total Sales')
            ws.cell(row, 2, self.total_sales_label.text().split(': ')[1])
            row += 1
            ws.cell(row, 1, 'Total Transactions')
            ws.cell(row, 2, self.total_transactions_label.text().split(': ')[1])
            row += 1
            ws.cell(row, 1, 'Total Cost')
            ws.cell(row, 2, self.total_cost_label.text().split(': ')[1])
            row += 1
            ws.cell(row, 1, 'Total Profit')
            ws.cell(row, 2, self.total_profit_label.text().split(': ')[1])
            
            # Sales table
            row += 2
            if self.table.rowCount() > 0:
                headers = ['Sale ID', 'Date', 'Cashier', 'Items', 'Total', 'Profit']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row, col, header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                row += 1
                for table_row in range(self.table.rowCount()):
                    for col in range(self.table.columnCount()):
                        item = self.table.item(table_row, col)
                        ws.cell(row, col + 1, item.text() if item else '')
                    row += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filename)
            show_info_dialog(self, "Success", f"Report exported to Excel:\n{filename}")
        except ImportError:
            show_error_dialog(self, "Error", "openpyxl package not installed")
        except Exception as e:
            show_error_dialog(self, "Error", f"Failed to export Excel: {str(e)}")

